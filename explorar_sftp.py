import paramiko
import openpyxl
import re
from decouple import config

host = config('SFTP_HOST')
port = 22
username = config('SFTP_USER')
password = config('SFTP_PASSWORD')

client = paramiko.SSHClient()
client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

try:
    client.connect(host, port, username, password)
    sftp = client.open_sftp()
    directorio_raiz = '/mnt/almacenamiento_energy/energy/masivo/TteGuia'
    sftp.chdir(directorio_raiz)
    limite = 1
    subdirectorios = sftp.listdir()
    for directorio in subdirectorios:
        print(f"Explorando directorio {directorio}")
        #workbook = openpyxl.Workbook()
        #sheet = workbook.active
        #sheet.title = "archivos"
        subdirectorio = f"/mnt/almacenamiento_energy/energy/masivo/TteGuia/{directorio}"
        sftp.chdir(subdirectorio)
        archivos_subdirectorio = sftp.listdir_attr()
        #sheet.cell(row=1, column=1, value="Nombre")
        #sheet.cell(row=1, column=2, value="Tamaño")
        fila = 1
        for idx, archivo_attr in enumerate(archivos_subdirectorio, start=2):
            nombreArchivo = archivo_attr.filename
            tamaño_archivo_bytes = archivo_attr.st_size
            tamaño_archivo_mb = tamaño_archivo_bytes / (1024 * 1024)
            print(nombreArchivo)
            #if tamaño_archivo_mb > 3:
                #fila += 1
                #sheet.cell(row=fila, column=1, value=nombre_archivo)
                #sheet.cell(row=fila, column=2, value=tamaño_archivo_mb)
        #nombre_archivo_excel = f"/home/desarrollo/Escritorio/analisis_ftp/listado_archivos{directorio}.xlsx"
        #workbook.save(nombre_archivo_excel)"""
        
        if limite == 1:
            break
        limite += 1
finally:
    sftp.close()
    client.close()